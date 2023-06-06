from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = Flask(__name__)
wb = load_workbook("Crimedb.xlsx")
ws = wb.active

worksheetrow = 1

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    global worksheetrow

    crimetitle = request.form['crime-description']
    crimelocation = request.form['your-location']
    crimedangerous = "10"
    now = datetime.now()

    ws.cell(row=worksheetrow, column=1).value = crimetitle
    ws.cell(row=worksheetrow, column=2).value = crimelocation
    ws.cell(row=worksheetrow, column=3).value = now
    worksheetrow += 1
    wb.save("Crimedb.xlsx")

    return render_template('submit.html', crimetitle=crimetitle, crimelocation=crimelocation, crimedangerous=crimedangerous)

@app.route('/receive')
def receive():
    workbook = load_workbook('Crimedb.xlsx')
    worksheet = workbook.active

    rows = []

    for row in worksheet.iter_rows(values_only=True):
        column1_value = row[0]
        column2_value = row[1]
        column3_value = row[2]

        rows.append((column1_value, column2_value, column3_value))

    return render_template('result.html', rows=rows)

if __name__ == '__main__':
    app.run(debug=True)
